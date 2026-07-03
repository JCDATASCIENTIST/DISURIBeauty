#!/usr/bin/env python3
"""Build the corrected GS1 US Data Hub import file.

Reads the original GS1 export plus the product master and produces an xlsx
with the same columns as the export, where:
  - Product Description = canonical_name from the master (clean short format)
  - SKU = kept if already filled in GS1, else the master internal_code
  - 3 new rows are appended for the skincare bundles with the next available
    GTIN-12s in the 0850066107 prefix block (check digits computed)

Archived GS1 rows are passed through untouched.

Outputs:
  DISURI Beauty/product-master/gs1-import-corrected.xlsx
  DISURI Beauty/product-master/gs1-upload-checklist.md

Usage:
  python3 scripts/build-gs1-import.py
"""
import csv
import os
import sys

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PM_DIR = os.path.join(REPO_ROOT, 'DISURI Beauty', 'product-master')
GS1_XLSX = os.path.join(PM_DIR, 'source', 'gs1-export-2026-07-02.xlsx')
MASTER_CSV = os.path.join(PM_DIR, 'product-master.csv')
OUT_XLSX = os.path.join(PM_DIR, 'gs1-import-corrected.xlsx')
OUT_MD = os.path.join(PM_DIR, 'gs1-upload-checklist.md')


def upc_check_digit(base11):
    digits = [int(c) for c in base11]
    odd = sum(digits[0::2])
    even = sum(digits[1::2])
    return str((10 - (odd * 3 + even) % 10) % 10)


def next_gtins(last_gtin12, count):
    base = int(last_gtin12[:11])
    out = []
    for i in range(1, count + 1):
        b = str(base + i).zfill(11)
        out.append(b + upc_check_digit(b))
    return out


# Registered manually in Data Hub 2026-07-03 with exactly these GTINs
# (850066107362/379/386) — do NOT append them to the import file again.
BUNDLES_REGISTERED = True

# GTINs archived in Data Hub 2026-07-03 (duplicates); excluded from the
# import file so the upload cannot error on or reactivate them.
SKIP_GTINS = {'850066107188', '850066107195'}

# The three skincare bundles (kept for the record; see BUNDLES_REGISTERED).
BUNDLES = [
    {
        'handle': 'the-complete-disuri-system',
        'description': 'DISURI Beauty The Complete DISURI System - 3-Step Korean Skincare Routine',
        'sku': 'DIS-SYS-COMPLETE',
    },
    {
        'handle': 'the-anti-aging-power-duo',
        'description': 'DISURI Beauty The Anti-Aging Power Duo - Korean Skincare Set',
        'sku': 'DIS-SYS-AADUO',
    },
    {
        # 850066107386 was registered as Barrier Rescue in the 2026-07-02
        # upload, then edited in Data Hub to become The Glass Skin Starter
        # (product line discontinued before ever trading).
        'handle': 'the-glass-skin-starter',
        'description': 'DISURI Beauty The Glass Skin Starter - 2-Step Korean Skincare Set',
        'sku': 'DIS-SYS-GLASS',
    },
]


def load_master():
    with open(MASTER_CSV, newline='') as f:
        return {r['gtin12']: r for r in csv.DictReader(f)}


def main():
    master = load_master()
    wb = openpyxl.load_workbook(GS1_XLSX)
    ws = wb['ExportAllProducts']
    hdr = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(hdr)}

    changed = 0
    sku_filled = 0
    skipped_archived = 0
    max_gtin = ''

    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        upc = str(row[col['GTIN-12 (U.P.C.)'] - 1].value or '').strip()
        if not upc:
            continue
        if upc > max_gtin:
            max_gtin = upc
        if upc in SKIP_GTINS:
            rows_to_delete.append(row[0].row)
            continue
        status = str(row[col['Status Label'] - 1].value or '').strip()
        if status == 'Archived':
            skipped_archived += 1
            continue
        m = master.get(upc)
        if not m:
            continue
        desc_cell = row[col['Product Description'] - 1]
        if (desc_cell.value or '').strip() != m['canonical_name']:
            desc_cell.value = m['canonical_name']
            changed += 1
        sku_cell = row[col['SKU'] - 1]
        if not str(sku_cell.value or '').strip() and m['internal_code']:
            sku_cell.value = m['internal_code']
            sku_filled += 1

    # Drop rows for GTINs archived directly in Data Hub (SKIP_GTINS).
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    # Append bundle rows only if they are not already registered in Data Hub.
    new_gtins = next_gtins(max_gtin, len(BUNDLES))
    if not BUNDLES_REGISTERED:
        for bundle, gtin in zip(BUNDLES, new_gtins):
            vals = [''] * len(hdr)
            vals[col['GS1 Company Prefix'] - 1] = '0850066107'
            vals[col['GTIN'] - 1] = '00' + gtin
            vals[col['GTIN-12 (U.P.C.)'] - 1] = gtin
            vals[col['Brand Name'] - 1] = 'DISURI BEAUTY'
            vals[col['Brand 1 Language'] - 1] = 'en'
            vals[col['Product Description'] - 1] = bundle['description']
            vals[col['Desc 1 Language'] - 1] = 'en'
            vals[col['Product Industry'] - 1] = 'General'
            vals[col['Packaging Level'] - 1] = 'Each'
            vals[col['Is Variable'] - 1] = 'N'
            vals[col['Is Purchasable'] - 1] = 'Y'
            vals[col['Status Label'] - 1] = 'In Use'
            vals[col['SKU'] - 1] = bundle['sku']
            vals[col['GPC Brick'] - 1] = '99999999 - Temporary Classification'
            vals[col['Target Markets'] - 1] = 'US'
            ws.append(vals)

    wb.save(OUT_XLSX)

    checklist = [
        '# GS1 Data Hub upload checklist',
        '',
        f'File: `gs1-import-corrected.xlsx` (generated by scripts/build-gs1-import.py)',
        '',
        f'- Descriptions corrected: {changed}',
        f'- Blank SKU fields filled with internal codes: {sku_filled}',
        f'- Archived rows left untouched: {skipped_archived}',
        f'- Data Hub-archived duplicate rows excluded from file: {len(rows_to_delete)}',
        '- Bundles registered manually in Data Hub 2026-07-03 (not in this file):'
        if BUNDLES_REGISTERED else
        '- New bundle GTINs proposed (verify Data Hub assigns these exact numbers):',
    ]
    for bundle, gtin in zip(BUNDLES, new_gtins):
        checklist.append(f'  - `{gtin}` -> {bundle["description"]} (SKU {bundle["sku"]})')
    checklist += [
        '',
        '## Steps',
        '',
        '1. Log in to GS1 US Data Hub -> Product -> Import.',
        '2. Upload `gs1-import-corrected.xlsx` (or paste corrections manually for',
        '   small batches — Data Hub import requires their template; if the upload',
        '   is rejected, export their blank template and map these columns over).',
        '3. Confirm the 3 new bundle GTINs; if Data Hub assigns different numbers,',
        '   record them in `overrides.csv` and update `HANDLE_TO_GTIN` in',
        '   `scripts/shopify-name-sku-sync.mjs`.',
        '4. After the bundles have GTINs, run:',
        '   `node scripts/shopify-name-sku-sync.mjs --export` then `--dry-run` then',
        '   `--allow-mutations` to push the bundle UPCs into Shopify SKU/barcode.',
        '5. Resolve the duplicates flagged in `conflict-review.md` (Los Angeles,',
        '   Firming Essence, Ultra Plumping Mango) by archiving the extra GTIN in',
        '   Data Hub.',
        '',
    ]
    with open(OUT_MD, 'w') as f:
        f.write('\n'.join(checklist))

    print(f'Wrote {OUT_XLSX}')
    print(f'  descriptions corrected: {changed}')
    print(f'  SKUs filled: {sku_filled}')
    if BUNDLES_REGISTERED:
        print('  bundles: already registered in Data Hub, not in file')
    else:
        print(f'  bundles appended: {[g for g in new_gtins]}')
    print(f'Checklist -> {OUT_MD}')


if __name__ == '__main__':
    sys.exit(main())
