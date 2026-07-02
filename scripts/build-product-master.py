#!/usr/bin/env python3
"""Build the DISURI canonical product master from the GS1 export and the
June 2024 Product Line sheet.

Sources (snapshotted in the repo):
  DISURI Beauty/product-master/source/gs1-export-2026-07-02.xlsx
  DISURI Beauty/product-master/source/disuri-product-line-2024-06-18.xlsx
  DISURI Beauty/product-master/source/shopify-catalog.json   (optional, from
      scripts/shopify-name-sku-sync.mjs --export)
  DISURI Beauty/product-master/overrides.csv                 (optional, manual
      shade/name overrides keyed by gtin12)

Outputs:
  DISURI Beauty/product-master/product-master.csv
  DISURI Beauty/product-master/conflict-review.md

Rules (agreed 2026-07-02):
  - GTINs from GS1 are the source of truth.
  - Shade names: GS1 wins by default; overrides.csv wins over everything.
  - Canonical name format ("clean short"):
      DISURI Beauty <Product Line> - <Shade> (<Code>)
    Code/shade parts are dropped when a product has none (skincare, bundles).

Usage:
  python3 scripts/build-product-master.py
"""
import csv
import json
import os
import re
import sys
from collections import OrderedDict

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PM_DIR = os.path.join(REPO_ROOT, 'DISURI Beauty', 'product-master')
SRC_DIR = os.path.join(PM_DIR, 'source')
GS1_XLSX = os.path.join(SRC_DIR, 'gs1-export-2026-07-02.xlsx')
SHEET_XLSX = os.path.join(SRC_DIR, 'disuri-product-line-2024-06-18.xlsx')
SHOPIFY_JSON = os.path.join(SRC_DIR, 'shopify-catalog.json')
OVERRIDES_CSV = os.path.join(PM_DIR, 'overrides.csv')
MASTER_CSV = os.path.join(PM_DIR, 'product-master.csv')
CONFLICTS_MD = os.path.join(PM_DIR, 'conflict-review.md')

MARKETING_JUNK = [
    ', Ultra Shiny, Non-Sticky, Vegan & Cruelty-Free Formula',
    ', Non-Sticky, Vegan & Cruelty-Free Formula',
    ' Ultra Shiny, Non-Sticky, Vegan & Cruelty-Free Formula',
    'Vegan & Cruelty-Free ',
]

# Canonical product-line labels, keyed by a matching phrase found in raw names.
# Order matters: first match wins (more specific phrases first).
LINE_MAP = [
    ('ultra plumping blast lip gloss', 'Ultra Plumping Blast Lip Gloss'),
    ('plumping blast lip gloss', 'Plumping Blast Lip Gloss'),
    ('lip plumping oil', 'Lip Plumping Oil'),
    ('shimmer blast lip gloss', 'Shimmer Blast Lip Gloss'),
    ('shimmer lip gloss', 'Shimmer Blast Lip Gloss'),
    ('glossy blast lip gloss', 'Glossy Blast Lip Gloss'),
    ('2-in-1 liquid lipstick lip liner kit', '2-in-1 Liquid Lipstick & Lip Liner Kit'),
    ('2 in 1 liquid lipstick lip liner kit', '2-in-1 Liquid Lipstick & Lip Liner Kit'),
    ('matte bullet lipstick', 'Matte Bullet Lipstick'),
    ('bullet lipstick', 'Matte Bullet Lipstick'),
    ('liquid lipstick', 'Liquid Lipstick'),
    ('color pop lip liner', 'Color Pop Lip Liner'),
    ('lip mask and lip scrub set', 'Lip Mask & Scrub Set'),
    ('lip mask & scrub set', 'Lip Mask & Scrub Set'),
    ('nail polish', 'Nail Polish'),
]

# Skincare / single-SKU products where the whole cleaned name is canonical.
STANDALONE = [
    ('hyaluronic acid essential intensive cream', 'Hyaluronic Acid Essential Intensive Cream'),
    ('triple collagen firming toner', 'Triple Collagen Firming Toner'),
    ('triple collagen firming cream', 'Triple Collagen Firming Cream'),
    ('triple collagen firming eye cream', 'Triple Collagen Firming Eye Cream'),
    ('triple collagen firming essence', 'Triple Collagen Firming Essence'),
    ('triple collagen firming cleansing foam', 'Triple Collagen Firming Cleansing Foam'),
    ('snail mucin cream', 'Snail Mucin Cream'),
    ('eyebrow', 'Brow Boost Eyebrow Growth Serum'),
    ('brow boost', 'Brow Boost Eyebrow Growth Serum'),
    ('lash max growth serum', 'Lash Max Lash Growth Serum'),
    ('eyelash growth', 'Lash Max Lash Growth Serum'),
    ('mascara', 'Lash Envy Fearless Mascara'),
]

BUNDLE_HINTS = ['collection', 'pack', 'complete', 'duo', 'system']

# Product lines Joel no longer sells (archived in Shopify 2026-07-02).
# GS1 GTINs stay registered; names still get corrected in the import file.
DISCONTINUED_LINES = {
    'Lip Plumping Oil',
    'Plumping Blast Lip Gloss',
    'Ultra Plumping Blast Lip Gloss',
}

# GTINs registered in GS1 Data Hub after the export snapshot (2026-07-02
# upload). 850066107386 was originally Barrier Rescue but was edited in Data
# Hub to become The Glass Skin Starter (never traded, so GS1 allows the edit).
EXTRA_PRODUCTS = [
    {
        'gtin12': '850066107362',
        'code': 'DIS-SYS-COMPLETE',
        'line': 'The Complete DISURI System - 3-Step Korean Skincare Routine',
        'kind': 'bundle',
    },
    {
        'gtin12': '850066107379',
        'code': 'DIS-SYS-AADUO',
        'line': 'The Anti-Aging Power Duo - Korean Skincare Set',
        'kind': 'bundle',
    },
    {
        'gtin12': '850066107386',
        'code': 'DIS-SYS-GLASS',
        'line': 'The Glass Skin Starter - 2-Step Korean Skincare Set',
        'kind': 'bundle',
    },
]


def norm(s):
    if s is None:
        return ''
    return ' '.join(str(s).split()).strip()


def clean_raw(name):
    s = norm(name)
    for junk in MARKETING_JUNK:
        s = s.replace(junk, '')
    s = s.replace('\u2013', '-').replace('\u2019', "'").replace('\u2018', "'")
    s = s.replace('\u201c', '"').replace('\u201d', '"')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def detect_line(s):
    low = s.lower()
    for phrase, label in LINE_MAP:
        if phrase in low:
            return label, 'shade'
    for phrase, label in STANDALONE:
        if phrase in low:
            return label, 'standalone'
    if any(h in low for h in BUNDLE_HINTS):
        return None, 'bundle'
    return None, 'unknown'


CODE_RE = r'#?[A-Za-z]{0,3}\d+[A-Za-z]?'


def split_shade_code(shade):
    """Strip a trailing ': CODE' / ':CODE' left inside a shade string."""
    m = re.match(rf'^(.*?)\s*:\s*({CODE_RE})\s*$', shade)
    if m:
        return m.group(1).strip(), m.group(2)
    return shade, ''


def parse_name(raw):
    """Return dict(line, shade, code, kind, needs_review)."""
    s = clean_raw(raw)
    line, kind = detect_line(s)
    out = {'line': line, 'shade': '', 'code': '', 'kind': kind, 'needs_review': False}

    if kind == 'standalone':
        return out

    if kind == 'bundle':
        # Bundles keep their given name minus brand prefix.
        name = re.sub(r'^disuri\s+beauty\s+', '', s, flags=re.I).strip()
        out['line'] = name
        return out

    if kind == 'unknown':
        out['needs_review'] = True
        out['line'] = re.sub(r'^disuri\s+beauty\s+', '', s, flags=re.I).strip()
        return out

    # kind == 'shade': try to pull the shade and internal code out.
    # Pattern A: quoted segment  "Shade: CODE"  or  "Shade" (code optional,
    # closing quote optional because GS1 data is often truncated).
    m = re.search(r'"\s*([^":]+?)\s*(?::\s*(#?[A-Za-z]{0,3}\d+[A-Za-z]?))?\s*(?:"|$)', s)
    if m:
        out['shade'] = m.group(1).strip().rstrip(',').strip()
        out['code'] = (m.group(2) or '').strip()
        return out

    # Pattern B: Color Pop Lip Liner — shade comes first: "Berry Temptation Color Pop Lip Liner | ..."
    m = re.match(r'^(.*?)\s+color pop lip liner\b', s, flags=re.I)
    if m and m.group(1) and not m.group(1).lower().startswith('disuri'):
        out['shade'] = m.group(1).strip()
        return out

    # Pattern C: bullet lipstick "... - Shade Name #01"
    m = re.search(r'-\s*([^-]+?)\s*(#\d+)\s*$', s)
    if m:
        out['shade'] = m.group(1).strip()
        out['code'] = m.group(2)
        return out

    # Pattern D: en-dash/pipe style "Lip Mask & Scrub Set – Vanilla S6 | ..."
    m = re.search(r'-\s*([A-Za-z][^|]*?)\s+(S\d+|M\d+)\s*\|', s)
    if m:
        out['shade'] = m.group(1).strip()
        out['code'] = m.group(2)
        return out

    # Pattern E: plain "<line> - Shade" (no quotes; may still carry ": CODE")
    m = re.search(r'-\s*([A-Za-z][^-|,]*?)\s*$', s)
    if m:
        shade, code = split_shade_code(m.group(1).strip())
        out['shade'] = shade
        out['code'] = code
        return out

    out['needs_review'] = True
    return out


def read_gs1():
    wb = openpyxl.load_workbook(GS1_XLSX, read_only=True)
    rows = list(wb['ExportAllProducts'].iter_rows(values_only=True))
    hdr = {h: i for i, h in enumerate(rows[0])}
    products = OrderedDict()
    for r in rows[1:]:
        upc = norm(r[hdr['GTIN-12 (U.P.C.)']])
        if not upc:
            continue
        products[upc] = {
            'gtin14': norm(r[hdr['GTIN']]),
            'desc': norm(r[hdr['Product Description']]),
            'sku_field': norm(r[hdr['SKU']]),
            'status': norm(r[hdr['Status Label']]),
        }
    return products


def read_sheet():
    wb = openpyxl.load_workbook(SHEET_XLSX, read_only=True, data_only=True)
    ws = wb['Product List']
    products = {}
    for r in ws.iter_rows(values_only=True):
        r = list(r) + [None] * 5
        sku = norm(r[2])
        if not re.fullmatch(r'\d{12}', sku):
            continue
        products[sku] = {
            'name': norm(r[1]),
            'category': norm(r[0]),
        }
    return products


def read_shopify():
    if not os.path.exists(SHOPIFY_JSON):
        return {}
    with open(SHOPIFY_JSON) as f:
        data = json.load(f)
    by_barcode = {}
    for p in data:
        for v in p.get('variants', []):
            bc = norm(v.get('barcode'))
            sku = norm(v.get('sku'))
            if re.fullmatch(r'\d{12}', bc):
                by_barcode[bc] = p['handle']
            elif re.fullmatch(r'\d{12}', sku):
                by_barcode[sku] = p['handle']
    return by_barcode


def read_overrides():
    if not os.path.exists(OVERRIDES_CSV):
        return {}
    overrides = {}
    with open(OVERRIDES_CSV, newline='') as f:
        for row in csv.DictReader(f):
            gtin = norm(row.get('gtin12'))
            if gtin:
                overrides[gtin] = {k: norm(v) for k, v in row.items()}
    return overrides


def upc_check_ok(gtin12):
    if not re.fullmatch(r'\d{12}', gtin12):
        return False
    d = [int(c) for c in gtin12[:11]]
    return int(gtin12[-1]) == (10 - (sum(d[0::2]) * 3 + sum(d[1::2])) % 10) % 10


def canonical_name(line, shade, code, kind):
    if kind == 'bundle':
        return f'DISURI Beauty {line}' if not line.lower().startswith('disuri') else line
    if kind in ('standalone', 'unknown') or not shade:
        return f'DISURI Beauty {line}'
    if code:
        return f'DISURI Beauty {line} - {shade} ({code})'
    return f'DISURI Beauty {line} - {shade}'


def main():
    gs1 = read_gs1()
    sheet = read_sheet()
    shopify = read_shopify()
    overrides = read_overrides()

    rows = []
    conflicts = []
    review_rows = []

    all_gtins = list(gs1.keys()) + [u for u in sheet if u not in gs1]
    for gtin in all_gtins:
        g = gs1.get(gtin)
        s = sheet.get(gtin)
        gp = parse_name(g['desc']) if g else None
        sp = parse_name(s['name']) if s else None

        base = gp or sp
        line = base['line'] or (sp['line'] if sp and sp['line'] else '')
        kind = base['kind']

        gs1_shade = gp['shade'] if gp else ''
        sheet_shade = sp['shade'] if sp else ''
        shade = gs1_shade or sheet_shade
        code = (gp['code'] if gp else '') or (sp['code'] if sp else '') or (g['sku_field'] if g else '')

        shade_conflict = bool(
            gs1_shade and sheet_shade and gs1_shade.lower() != sheet_shade.lower()
        )

        source_flags = []
        if not g:
            source_flags.append('sheet-only (no GS1 row)')
        if not s:
            source_flags.append('gs1-only (not in June 2024 sheet)')
        if line in DISCONTINUED_LINES:
            source_flags.append('discontinued (archived in Shopify 2026-07-02)')
        if not upc_check_ok(gtin):
            source_flags.append('INVALID GTIN (fails UPC check digit — do not use)')

        ov = overrides.get(gtin, {})
        if ov.get('shade'):
            shade = ov['shade']
        if ov.get('code'):
            code = ov['code']
        if ov.get('line'):
            line = ov['line']

        name = ov.get('canonical_name') or canonical_name(line, shade, code, kind)

        needs_review = (gp['needs_review'] if gp else False) or (not line)
        row = {
            'gtin12': gtin,
            'gtin14': g['gtin14'] if g else '',
            'internal_code': code,
            'product_line': line,
            'shade_name': shade,
            'canonical_name': name,
            'gs1_name_current': g['desc'] if g else '',
            'sheet_name': s['name'] if s else '',
            'sheet_category': s['category'] if s else '',
            'shopify_handle': shopify.get(gtin, ''),
            'status': g['status'] if g else 'NOT-IN-GS1',
            'shade_conflict': 'Y' if shade_conflict else '',
            'override_applied': 'Y' if ov else '',
            'needs_review': 'Y' if needs_review else '',
            'notes': '; '.join(source_flags),
        }
        rows.append(row)
        if shade_conflict and not ov.get('shade'):
            conflicts.append(row | {'gs1_shade': gs1_shade, 'sheet_shade': sheet_shade})
        if needs_review:
            review_rows.append(row)

    existing = {r['gtin12'] for r in rows}
    for extra in EXTRA_PRODUCTS:
        if extra['gtin12'] in existing:
            continue
        rows.append({
            'gtin12': extra['gtin12'],
            'gtin14': '00' + extra['gtin12'],
            'internal_code': extra['code'],
            'product_line': extra['line'],
            'shade_name': '',
            'canonical_name': canonical_name(extra['line'], '', extra['code'], extra['kind']),
            'gs1_name_current': '',
            'sheet_name': '',
            'sheet_category': '',
            'shopify_handle': shopify.get(extra['gtin12'], ''),
            'status': 'In Use',
            'shade_conflict': '',
            'override_applied': '',
            'needs_review': '',
            'notes': 'registered post-export (Data Hub upload 2026-07-02)',
        })

    os.makedirs(PM_DIR, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with open(MASTER_CSV, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    write_conflict_review(conflicts, review_rows, rows)

    n_conf = len(conflicts)
    n_rev = len(review_rows)
    print(f'product-master.csv: {len(rows)} rows -> {MASTER_CSV}')
    print(f'shade conflicts (GS1 default in effect): {n_conf}')
    print(f'rows flagged needs_review: {n_rev}')
    print(f'overrides applied: {sum(1 for r in rows if r["override_applied"])}')


def write_conflict_review(conflicts, review_rows, rows):
    lines = [
        '# Conflict review — DISURI product master',
        '',
        'Generated by `scripts/build-product-master.py`. To override a row, add it to',
        '`overrides.csv` (columns: `gtin12,shade,code,line,canonical_name,note` — only',
        '`gtin12` plus the fields you want to change) and re-run the script.',
        '',
        '## Shade-name conflicts (GS1 vs June 2024 sheet)',
        '',
        'GS1 shade is the default and is what the master currently uses.',
        'Mark any row where the **sheet** name is actually correct.',
        '',
        '| GTIN-12 | Product line | GS1 shade (DEFAULT) | Sheet shade | Use sheet instead? |',
        '|---|---|---|---|---|',
    ]
    for c in conflicts:
        lines.append(
            f"| {c['gtin12']} | {c['product_line']} | **{c['gs1_shade']}** | {c['sheet_shade']} | [ ] |"
        )

    lines += [
        '',
        '## Known duplicates / data issues',
        '',
        '- **"Los Angeles: 37" has two GTINs**: `850064314824` (June 2024 sheet) and',
        '  `850066107171` (GS1). RESOLVED: `850064314824` fails UPC check-digit',
        '  validation — it was a typo in the sheet. `850066107171` is the real GTIN',
        '  (already live on Shopify). No GS1 action needed.',
        '- **Triple Collagen Firming Essence has two GTINs in GS1**: `850064676358`',
        '  and `850066107188`. Verify one is not a mistake; archive the extra.',
        '- **Ultra Plumping Blast Lip Gloss - Mango (8) has two GTINs in GS1**:',
        '  `850064676181` and `850066107195`. Verify; archive the extra.',
        '- **`850066107317` "Shimmer Lip Gloss - Sunset Glow"** has no internal code',
        '  and is not in the June 2024 sheet — confirm which shade/code this is.',
        '',
        '## Rows needing manual review (parser could not fully classify)',
        '',
        '| GTIN-12 | GS1 name | Parsed as |',
        '|---|---|---|',
    ]
    for r in review_rows:
        lines.append(f"| {r['gtin12']} | {r['gs1_name_current'][:80]} | {r['canonical_name'][:60]} |")

    lines += [
        '',
        '## Cross-source coverage',
        '',
        f'- Total GTINs in master: {len(rows)}',
        f"- GS1-only (no June 2024 sheet row): {sum(1 for r in rows if 'gs1-only' in r['notes'])}",
        f"- Sheet-only (no GS1 row): {sum(1 for r in rows if 'sheet-only' in r['notes'])}",
        '',
    ]
    with open(CONFLICTS_MD, 'w') as f:
        f.write('\n'.join(lines))


if __name__ == '__main__':
    sys.exit(main())
